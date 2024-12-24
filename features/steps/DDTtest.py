import openpyxl
from behave import *

from features.pages.AccountPage import AccountPage
from features.pages.HomePage import HomePage
from features.pages.LoginPage import LoginPage


@given(u'I got navigated to login')
def step_impl(context):
    context.homepage = HomePage(context.driver)
    context.homepage.click_on_account_option()
    context.loginpage = context.homepage.click_on_Login_option()


@when(u'I am entering the email address as "{email}" and valid password as "{password}" into the fields')
def step_impl(context, email, password):
    context.email = email
    context.loginpage.enter_the_email(email)
    context.loginpage.enter_the_password(password)


@when(u'I click on the login')
def step_impl(context):
    context.accountpage = context.loginpage.click_on_login()


@then(u'Get into the website if the credientials or correct')
def step_impl(context):
    workbook = openpyxl.load_workbook(r"C:\Users\Admin\PycharmProjects\BDDBehave\Book1.xlsx")
    sheet = workbook.active
    no_of_rows = sheet.max_row
    no_of_cols = sheet.max_column
    if context.accountpage.edit_yout_account_asserti():
        for i in range(1,no_of_rows+1):
            if sheet.cell(i,1).value == context.email:
                sheet.cell(i,3).value = "Passed Authentication"
    else:
        for i in range(1, no_of_rows + 1):
            if sheet.cell(i, 1).value == context.email:
                sheet.cell(i, 3).value = "Failed Authentication"

    workbook.save(r"C:\Users\Admin\PycharmProjects\BDDBehave\Book1.xlsx")
